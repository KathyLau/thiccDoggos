import sqlite3
from time import gmtime, localtime, strftime
from datetime import datetime
#the location of the db
f = "data/users.db"


def create_data():
    from openpyxl import load_workbook
    db=sqlite3.connect(f)
    c=db.cursor()
    wb = load_workbook('KLsched.xlsx')
    ws = wb['content']
    #print result
    #c.execute('CREATE TABLE usersA(osis STRING, pd INTEGER);')
    #c.execute('CREATE TABLE usersB(osis STRING, pd INTEGER);')
    #c.execute('CREATE TABLE banlist(osis STRING);')
    for row in ws.rows[1:ws.max_row]:
        count = 1
        for cell in row[5:15]:
            if str(cell.value)[:1] in 'Z' or cell.value==None:
                #print 'hi'
                c.execute('INSERT INTO usersA VALUES("%s", "%s");' % (str(row[1].value)  , str(count) ))
                c.execute('INSERT INTO usersB VALUES("%s", "%s");' % (str(row[1].value)  , str(count) ))
            if str(cell.value)[:2] in ['PE', 'PP', 'PO', 'PF', 'PW']:
                cell = str(cell.value)
                if cell[cell.find('/')-1]=="A":
                    c.execute('INSERT INTO usersA VALUES("%s", "%s");' % (str(row[1].value)  , str(count) ))
                if cell[cell.find('/')-1]=="B":
                    c.execute('INSERT INTO usersB VALUES("%s", "%s");' % (str(row[1].value)  , str(count) ))
            count += 1
    db.commit()
    db.close()

#create_data()
def count(s):
    pd = getPeriod(s)
    db=sqlite3.connect(f)
    c=db.cursor()
    date = strftime("%m-%d-%y", localtime())
    q="SELECT * FROM log WHERE date==" + "'" + str(date) + "' AND pd ==" + str(pd) + ';'
    c.execute(q)
    count = c.fetchall()
    count = len(count) + 1
    return count


def testadd(osis, day, s):
    pd = getPeriod(s)
    db=sqlite3.connect(f)
    c=db.cursor()
    ret = False
    co = count(s)

    #print banlist()
    if int(osis) in banlist():
        return [False, co]
    if day=="A":
        #print "SELECT * FROM usersA WHERE osis==" + str(osis) + " AND pd ==" + str(pd) + ';'
        c.execute("SELECT * FROM usersA WHERE osis==" + str(osis) + " AND pd ==" + str(pd) + ';')
    else:
        c.execute("SELECT * FROM usersB WHERE osis==" + str(osis) + " AND pd ==" + str(pd) + ';')
    ca = c.fetchall()

    if len(ca)> 0:
        ret = True
        date = strftime("%m-%d-%y", localtime())
        c.execute('INSERT INTO log VALUES("%s", "%s", "%s", "%s");' % (date, pd, osis, co))
        db.commit()
    return [ret, co]


#print testadd('123457051', '4', 'A')

def retAll():
    db=sqlite3.connect(f)
    c=db.cursor()
    date = strftime("%m-%d-%y", localtime())
    c.execute("SELECT * FROM log WHERE date ==" + "'" + str(date) + "'")
    ca = c.fetchall()
    return ca
#print retAll()

def retSearch(day, p):
    db=sqlite3.connect(f)
    c=db.cursor()
    if p == 'allday':
        c.execute("SELECT * FROM log WHERE date ==" + "'" + str(day) + "'");
    else:
        c.execute("SELECT * FROM log WHERE date ==" + "'" + str(day) + "' AND pd ==" + str(p) + ';')
    ca = c.fetchall()
    return ca
#print retAll()

def getPeriod(s):
    Rperiod = [
        [460,521],
        [520,566],
        [565,615],
        [614,661],
        [660,707],
        [706,753],
        [752,799],
        [798,845],
        [844,890],
        [889,935]
    ]

    Hperiod = [
        [460,520],
        [519,565],
        [564,609],
        [608,669],
        [668,714],
        [713,758],
        [757,802],
        [801,846],
        [845,890],
        [889,935]#,
        #[934,1440],

    ]

    time = datetime.now()
    hour = time.strftime('%H')
    minute = time.strftime('%M')
    Tminute = 60 * int(hour) + int(minute)
    i = 0
    #print Tminute
    if s =='r':
        llist = Rperiod
    else:
        llist = Hperiod
    #print "\n\n\n\n\n", llist
    while i < len(llist):
        #print Rperiod[i]
        if Tminute >= llist[i][0] and Tminute <= llist[i][1]:
            return i+1
        else:
            i+=1
    return 10
#print getPeriod()

def ban(osis):
    db=sqlite3.connect(f)
    c=db.cursor()
    c.execute('INSERT INTO banlist VALUES("%s");' % (osis))
    db.commit()
    return osis

def banlist():
    db=sqlite3.connect(f)
    c=db.cursor()
    c.execute('SELECT * from banlist;')
    ca = c.fetchall()
    db.commit()
    return [item[0] for item in ca]

def removeban(osis):
    db=sqlite3.connect(f)
    c=db.cursor()
    c.execute('DELETE from banlist where osis ==' + osis + ';')
    db.commit()
    return osis
